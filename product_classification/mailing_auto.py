from email.message import EmailMessage
from smtplib import SMTP_SSL
from pathlib import Path
from openpyxl import load_workbook

class SendEmail:
    #init은 인스턴스 객체 초기화
    def __init__(self, id, pw, df):
        # 로그인 계정/pw
        self.id = id
        self.pw = pw  # 앱비밀번호16자리
        # 엑셀파일에서 가져온 정보를 활용해 함수 반복 실행
        self.df = df

    def send_email(self):
        #email_list 실행
        wb = load_workbook(self.df, data_only=True)
        ws = wb.active

        #read_email_list, for문 안에서 email 발송
        for row in ws.iter_rows():
            recipient = row[0].value
            print(recipient)
            title = row[1].value
            text = row[2].value
            attachment = row[3].value

            # 템플릿 생성
            msg = EmailMessage()

            # 보내는 사람 / 받는 사람 / 제목 입력
            msg["From"] = self.id
            msg["To"] = recipient.split(",")
            msg["Subject"] = title

            # 본문 구성
            msg.set_content(text)

            # 파일 첨부
            if attachment:
                filenm = Path(attachment).name
                with open('./data/'+attachment, 'rb') as f:
                    msg.add_attachment(f.read(), maintype='application', subtype='octet-stream', filename=filenm)

            # 보내는 사람 로그인 및 smtp 서버로 발송
            with SMTP_SSL("smtp.gmail.com", 465) as smtp:
                smtp.login(self.id, self.pw)
                smtp.send_message(msg)
            # 완료 메시지
            print("발송 성공")

#인스턴스 생성
es = SendEmail('**', '**', 'email_list.xlsx')

#메소드 호출
es.send_email()

